import streamlit as st
import requests
import pandas as pd
from datetime import datetime
import time
from urllib.parse import urlparse
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import google.generativeai as genai
from PIL import Image

# ▼▼▼ 設定エリア (楽天) ▼▼▼
APP_ID = '1052224946268447244' 
REVIEW_RATE = 0.08  
PRICE_UPLIFT = 1.2  

# --- ページ設定 ---
st.set_page_config(page_title="EC運営支援ツール Suite Pro", page_icon="🛍️", layout="wide")

# --- CSSスタイル ---
st.markdown("""
<style>
    .main { padding-top: 2rem; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #BF0000; color: white; font-weight: bold; }
    .stDownloadButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #008000; color: white; }
    textarea { font-family: monospace; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# 共通・ロジック関数群 (楽天)
# ==========================================

def get_item_key_from_url(url):
    try:
        parsed = urlparse(url)
        path_parts = [p for p in parsed.path.split('/') if p]
        if len(path_parts) >= 2: return path_parts[-1]
        return url
    except: return url

def calculate_metrics(item, uplift, rate):
    price = item['itemPrice']
    review_count = item['reviewCount']
    item_name = item['itemName']
    catch_copy = item.get('catchcopy', '')
    
    adj_price = int(price * uplift)
    total_sales_vol = int(review_count / rate)
    total_sales_amt = total_sales_vol * adj_price
    
    full_text = (item_name + catch_copy).replace(" ", "")
    coupon_flg = "-"
    if any(x in full_text for x in ["クーポン", "OFF", "値引", "SALE"]):
        coupon_flg = "有"
    
    return {
        "商品名": item_name, "価格": price, "ポイント倍率": item['pointRate'],
        "クーポン有無": coupon_flg, "レビュー総数": review_count,
        "推定累積販売数": total_sales_vol, "推定累積売上": total_sales_amt,
        "ショップ名": item['shopName'], "ショップコード": item['shopCode'],
        "商品URL": item['itemUrl'], "ジャンルID": item['genreId']
    }

def search_items(query, limit=10):
    url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20170706"
    if "http" in query:
        keyword = get_item_key_from_url(query)
        search_type = "URL検索"
    elif query.isdigit() and len(query) > 7:
        keyword = query
        search_type = "JAN検索"
    else:
        keyword = query
        search_type = "ワード検索"

    params = {"applicationId": APP_ID, "keyword": keyword, "hits": limit, "sort": "-reviewCount", "availability": 1}
    try:
        time.sleep(0.5)
        res = requests.get(url, params=params, timeout=10)
        data = res.json()
        results = []
        if 'Items' in data:
            for w in data['Items']:
                metrics = calculate_metrics(w['Item'], PRICE_UPLIFT, REVIEW_RATE)
                metrics['検索条件'] = query
                metrics['検索タイプ'] = search_type
                results.append(metrics)
        return results
    except: return []

def get_shop_top_items(shop_code, shop_name, limit=30):
    url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20170706"
    params = {"applicationId": APP_ID, "shopCode": shop_code, "hits": limit, "sort": "-reviewCount", "availability": 1}
    try:
        time.sleep(0.5)
        res = requests.get(url, params=params, timeout=10)
        data = res.json()
        results = []
        if 'Items' in data:
            for w in data['Items']:
                metrics = calculate_metrics(w['Item'], PRICE_UPLIFT, REVIEW_RATE)
                metrics['対象店舗'] = shop_name
                results.append(metrics)
        return results
    except: return []

def get_current_price_for_rpp(item_manage_number, shop_code):
    url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20170706"
    keyword = str(item_manage_number).strip()
    if ":" in keyword:
        keyword = keyword.split(":")[-1]

    params = {
        "applicationId": APP_ID, "shopCode": shop_code,
        "keyword": keyword, "hits": 1
    }
    
    try:
        res = requests.get(url, params=params, timeout=5)
        data = res.json()
        
        if res.status_code == 200:
            if 'Items' in data and len(data['Items']) > 0:
                return data['Items'][0]['Item']['itemPrice'], "成功"
            else:
                return None, "該当なし"
        else:
            return None, f"APIエラー({res.status_code})"
    except Exception as e:
        return None, "通信エラー"

def clean_number(val, default_val=0):
    if pd.isna(val): return default_val
    s_val = str(val).replace(',', '').replace('円', '').replace('%', '').strip()
    if s_val == '' or s_val.lower() == 'nan': return default_val
    try:
        return float(s_val)
    except:
        return default_val

def format_worksheet(worksheet):
    left_align = Alignment(horizontal='left', vertical='center')
    fill_color = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    hyperlink_font = Font(color="0000FF", underline="single")
    
    num_cols = ["価格", "レビュー総数", "推定累積販売数", "推定累積売上", 
                "現在価格", "入札単価", "推奨入札単価", "商品CPC", "クリック数(合計)", 
                "実績額(合計)", "CPC実績(合計)", "売上金額(合計720時間)", "売上件数(合計720時間)", "注文獲得単価(合計720時間)"]
    
    for row in worksheet.iter_rows():
        worksheet.row_dimensions[row[0].row].height = 25
        for cell in row:
            cell.alignment = left_align
            if cell.row == 1:
                cell.fill = fill_color
                continue
            
            header_val = worksheet.cell(row=1, column=cell.column).value
            if header_val in num_cols:
                cell.number_format = '#,##0'
            if header_val == "商品URL" and cell.value:
                cell.hyperlink = cell.value
                cell.font = hyperlink_font

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions
    for col in worksheet.columns:
        column = get_column_letter(col[0].column)
        worksheet.column_dimensions[column].width = 18

# ==========================================
# 共通・ロジック関数群 (ブログAI生成: Pro版)
# ==========================================
def generate_blog_content(api_key, image, keywords, tone):
    genai.configure(api_key=api_key)
    
    prompt = f"""
    あなたはプロのECサイト運営者兼ブロガーです。
    アップロードされた商品画像を見て、以下の条件でブログ記事を作成してください。

    【ターゲット・キーワード】
    {keywords}

    【文体のトーン】
    {tone}

    【出力要件】
    1. 記事のタイトルを作成してください。
    2. 記事の本文は、そのままWordPressやShopifyに貼り付けられる「HTML形式」で出力してください。
    3. <h2>, <h3>, <p>, <ul>, <li> などのタグを適切に使用し、読みやすくしてください。
    4. 画像の視覚的特徴（色、素材、雰囲気）を具体的に描写し、読者が商品をイメージできるようにしてください。
    5. SEOを意識し、キーワードを自然に盛り込んでください。
    6. 記事の最後には、購買意欲をそそるまとめを書いてください。
    
    【画像生成用プロンプト】
    記事の最後に、別途「この商品を魅力的なシーンで撮影したような画像をAIで作るための英語の指示文（Prompt）」を作成してください。
    （例: A photorealistic shot of a ceramic vase on a wooden table, sunlight streaming through a window, cozy scandinavian style, 8k resolution...）
    """

    # 優先順位: 
    # 1. gemini-1.5-pro (最高性能・今回のご希望)
    # 2. gemini-1.5-flash (高速版)
    # 3. gemini-pro-vision (旧・画像対応版)
    
    if image:
        models_to_try = ['gemini-1.5-pro', 'gemini-1.5-flash', 'gemini-pro-vision']
    else:
        models_to_try = ['gemini-1.5-pro', 'gemini-1.5-flash', 'gemini-pro']

    last_error = None
    
    for model_name in models_to_try:
        try:
            model = genai.GenerativeModel(model_name)
            if image:
                response = model.generate_content([prompt, image])
            else:
                response = model.generate_content(prompt)
            return response.text # 成功したら終了
            
        except Exception as e:
            last_error = e
            continue # 失敗したら次のモデルへ

    return f"エラー: 全てのAIモデルで生成に失敗しました。\n詳細: {last_error}"

# ==========================================
# メインアプリケーション
# ==========================================
def main():
    st.title("EC運営支援ツール Suite Pro")
    
    tab1, tab2, tab3 = st.tabs(["📊 楽天:競合分析", "💰 楽天:RPP改善", "📝 ブログ自動生成"])

    # -----------------------------------
    # Tab 1: 競合分析
    # -----------------------------------
    with tab1:
        st.subheader("競合・市場調査")
        st.markdown("調査したい **キーワード、JAN、URL** を入力してください。")
        input_text = st.text_area("検索リスト", height=150, placeholder="例:\n北欧 花瓶\n4968912801046", key="comp_input")
        
        if st.button("分析を開始する", key="comp_btn"):
            if not input_text.strip():
                st.warning("キーワードが入力されていません。")
            else:
                target_list = [{'query': line.strip()} for line in input_text.split('\n') if line.strip()]
                status_text = st.empty()
                progress_bar = st.progress(0)
                
                try:
                    sheet1_data = []
                    analyzed_shops = set()
                    
                    # Search
                    total = len(target_list)
                    for i, target in enumerate(target_list):
                        q = target['query']
                        status_text.text(f"検索中 ({i+1}/{total}): {q}")
                        items = search_items(q, limit=10)
                        sheet1_data.extend(items)
                        for item in items:
                            if item['ショップコード'] not in analyzed_shops:
                                analyzed_shops.add(item['ショップコード'])
                        progress_bar.progress(int((i+1) / total * 40))

                    # Shop Analysis
                    sheet2_data = []
                    total_shops = len(analyzed_shops)
                    status_text.text(f"店舗詳細分析中... (全{total_shops}店舗)")
                    shop_map = {row['ショップコード']: row['ショップ名'] for row in sheet1_data}
                    
                    for i, s_code in enumerate(analyzed_shops):
                        s_name = shop_map.get(s_code, "不明")
                        shop_items = get_shop_top_items(s_code, s_name, limit=30)
                        sheet2_data.extend(shop_items)
                        current_progress = 40 + int((i+1) / max(1, total_shops) * 60)
                        progress_bar.progress(min(100, current_progress))

                    status_text.text("Excel生成中...")
                    df1 = pd.DataFrame(sheet1_data)
                    df2 = pd.DataFrame(sheet2_data)
                    
                    output = io.BytesIO()
                    if not df1.empty: df1 = df1.sort_values(by='推定累積売上', ascending=False)
                    cols1 = ['検索タイプ', '検索条件', '商品名', '価格', 'レビュー総数', '推定累積販売数', '推定累積売上', 'ポイント倍率', 'クーポン有無', 'ショップ名', '商品URL']
                    df1 = df1.reindex(columns=cols1) if not df1.empty else pd.DataFrame()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        if not df1.empty:
                            df1.to_excel(writer, sheet_name='検索結果', index=False)
                            format_worksheet(writer.sheets['検索結果'])
                        if not df2.empty:
                            df2.to_excel(writer, sheet_name='店舗分析', index=False)
                            format_worksheet(writer.sheets['店舗分析'])
                    
                    progress_bar.progress(100)
                    status_text.success("分析完了！")
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
                    st.download_button(
                        label="📊 分析結果Excelをダウンロード",
                        data=output.getvalue(),
                        file_name=f"rakuten_analysis_{timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"エラーが発生しました: {e}")

    # -----------------------------------
    # Tab 2: RPP広告改善
    # -----------------------------------
    with tab2:
        st.subheader("RPP広告 CPC自動最適化")
        st.markdown("RMSのパフォーマンスレポートをアップロードしてください。")

        col1, col2 = st.columns(2)
        with col1:
            my_shop_code = st.text_input("自店舗ID (URLの英数字)", value="lykke-hygge", help="例: lykke-hygge")
        with col2:
            uploaded_file = st.file_uploader("RPP実績ファイル (CSV/Excel)", type=['csv', 'xlsx', 'xls'])

        with st.expander("詳細設定", expanded=True):
            c1, c2, c3, c4 = st.columns(4)
            target_roas = c1.number_input("目標ROAS (%)", min_value=100, value=400, step=50)
            min_cpc = c2.number_input("最低入札単価 (円)", min_value=10, value=25)
            max_cpc = c3.number_input("最高入札単価 (円)", min_value=10, value=100)
            skip_rows_num = c4.number_input("ヘッダー開始行", min_value=1, value=7)

        if st.button("価格取得＆改善実行", key="rpp_btn"):
            if not uploaded_file or not my_shop_code:
                st.error("ファイルと自店舗IDは必須です。")
            else:
                try:
                    df_rpp = None
                    skip_rows_count = skip_rows_num - 1 

                    if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
                        uploaded_file.seek(0)
                        try:
                            df_rpp = pd.read_excel(uploaded_file, skiprows=skip_rows_count)
                        except: pass
                    else:
                        encodings = ['shift_jis', 'cp932', 'utf-8', 'utf-8-sig']
                        for enc in encodings:
                            try:
                                uploaded_file.seek(0)
                                df_rpp = pd.read_csv(uploaded_file, encoding=enc, skiprows=skip_rows_count)
                                if len(df_rpp.columns) > 1: break
                            except: continue
                    
                    if df_rpp is None:
                        st.error(f"読み込み失敗。ヘッダー開始行({skip_rows_num}行目)の設定を確認してください。")
                        st.stop()
                    
                    req_cols = [
                        "商品管理番号", "入札単価", "CTR(%)", "商品CPC", "クリック数(合計)", 
                        "実績額(合計)", "CPC実績(合計)", "売上金額(合計720時間)", 
                        "売上件数(合計720時間)", "CVR(合計720時間)(%)", "ROAS(合計720時間)(%)", 
                        "注文獲得単価(合計720時間)"
                    ]
                    
                    if "商品管理番号" not in df_rpp.columns:
                        st.error(f"CSVの中に「商品管理番号」列が見つかりません。")
                        st.stop()
                    
                    st.write(f"データ件数: {len(df_rpp)}件")
                    progress_rpp = st.progress(0)
                    results_rpp = []
                    total_rows = len(df_rpp)
                    
                    for index, row in df_rpp.iterrows():
                        progress_rpp.progress((index + 1) / total_rows)
                        
                        item_manage_number = str(row.get("商品管理番号", "")).strip()
                        if not item_manage_number or item_manage_number.lower() == 'nan': continue
                        
                        current_bid = clean_number(row.get("入札単価"), 25)
                        actual_cpc = clean_number(row.get("CPC実績(合計)"), 25)
                        roas = clean_number(row.get("ROAS(合計720時間)(%)"), 0)
                        clicks = int(clean_number(row.get("クリック数(合計)"), 0))
                        
                        current_price, status_msg = get_current_price_for_rpp(item_manage_number, my_shop_code)
                        time.sleep(0.3)
                        
                        base_cpc = current_bid if current_bid > 0 else actual_cpc
                        new_bid = base_cpc
                        reason = "維持"
                        
                        if roas == 0 and clicks > 20:
                            new_bid = max(min_cpc, base_cpc - 10)
                            reason = "クリック過多・売上なし"
                        elif 0 < roas < target_roas:
                            new_bid = max(min_cpc, base_cpc - 5)
                            reason = "ROAS低・抑制"
                        elif roas > (target_roas + 200):
                            new_bid = min(max_cpc, base_cpc + 10)
                            reason = "ROAS好調・強化"
                        
                        row_data = {
                            "商品管理番号": item_manage_number,
                            "現在価格": current_price if current_price else "取得失敗",
                            "APIステータス": status_msg,
                            "推奨入札単価": int(new_bid),
                            "変更理由": reason
                        }
                        for col in req_cols:
                            if col != "商品管理番号":
                                row_data[col] = row.get(col, "")
                                
                        results_rpp.append(row_data)
                    
                    if not results_rpp:
                        st.warning("処理データなし")
                    else:
                        first_cols = ["商品管理番号", "現在価格", "推奨入札単価", "変更理由", "入札単価", "APIステータス"]
                        other_cols = [c for c in req_cols if c not in ["商品管理番号", "入札単価"]]
                        final_cols = first_cols + other_cols
                        
                        df_res = pd.DataFrame(results_rpp)
                        existing_cols = [c for c in final_cols if c in df_res.columns]
                        df_res = df_res[existing_cols]
                        
                        st.success("完了！")
                        st.dataframe(df_res)
                        
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_res.to_excel(writer, sheet_name='RPP改善案', index=False)
                            format_worksheet(writer.sheets['RPP改善案'])
                        
                        st.download_button(
                            label="推奨CPCリストをダウンロード (Excel)",
                            data=output.getvalue(),
                            file_name='rpp_optimized_v7.xlsx',
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                except Exception as e:
                    st.error(f"予期せぬエラー: {e}")

    # -----------------------------------
    # Tab 3: ブログ自動生成
    # -----------------------------------
    with tab3:
        st.subheader("📝 毎日ブログ自動生成くん")
        st.markdown("商品画像をアップするだけで、**HTML記事** と **画像生成用プロンプト** を作成します。")

        # 設定エリア
        with st.expander("設定 (Gemini API Key)", expanded=True):
            gemini_key = st.text_input("Google Gemini API Key", type="password", key="blog_gemini_key")

        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("1. 素材をセット")
            uploaded_img = st.file_uploader("商品画像をアップロード (必須)", type=['jpg', 'png', 'jpeg'], key="blog_img")
            
            keywords = st.text_area("キーワード・ターゲット", height=100, 
                                    placeholder="例：\n30代女性、北欧インテリア\n癒やし、ドライフラワー、プレゼント", key="blog_kw")
            
            tone = st.selectbox("記事の雰囲気", ["親しみやすい・共感", "高級感・プロフェッショナル", "シンプル・ミニマル", "情熱的・セールス強め"], key="blog_tone")

            if uploaded_img:
                st.image(uploaded_img, caption="対象の商品画像", use_column_width=True)

        with col2:
            st.subheader("2. 生成結果")
            
            if st.button("🚀 ブログ記事を生成する", key="blog_btn"):
                if not gemini_key:
                    st.error("設定エリアにGemini APIキーを入力してください。")
                elif not uploaded_img:
                    st.error("商品画像をアップロードしてください。")
                else:
                    try:
                        with st.spinner("画像を見て、記事を書いています...（約30秒）"):
                            # 画像データの準備
                            img = Image.open(uploaded_img)
                            
                            # AI実行
                            result_text = generate_blog_content(gemini_key, img, keywords, tone)
                            
                            st.success("生成完了！")
                            
                            # 結果の表示
                            st.text_area("HTMLコード (コピーしてブログに貼り付け)", value=result_text, height=600)
                            
                            st.info("💡 ヒント: 上記のテキストの最後にある『画像生成用プロンプト』をコピーして、MidjourneyやChatGPT(DALL-E)に貼り付けると、イメージ画像が作れます。")
                            
                    except Exception as e:
                        st.error(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main()
